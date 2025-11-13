"""Fetch test rows and replay them against the Boom Guru endpoint.

Supports sourcing rows from the database, a single image URL, or an Excel file.
"""
from __future__ import annotations

import argparse
import json
import logging
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Iterable, Sequence, Union
from uuid import uuid4

import requests

from src.config import MSSQL_DATABASE
from src.db import get_db_connection

LOGGER = logging.getLogger(__name__)
DEFAULT_ENDPOINT = "http://localhost:8361/boom_guru"
DEFAULT_OUTPUT_PATH = Path("boom_guru_test_responses.json")
DEFAULT_WEBHOOK_URL = "http://localhost:8092/webhook-receiver"
DEFAULT_ORDER_CANDIDATES: Sequence[str] = (
    "created_at",
    "created_on",
    "created_date",
    "inserted_at",
    "inserted_on",
    "updated_at",
    "request_date",
    "request_time",
    "id",
)
REQUIRED_FIELDS = ("image_url", "image_id", "serial_number", "webhook_url", "language")
OPTIONAL_FIELDS = ("form_id", "question_id")
TABLE_SCHEMA = "dbo"
TABLE_NAME = "BOOM_GURU_TEST"
FULL_TABLE_NAME = f"AIRPA.{TABLE_SCHEMA}.{TABLE_NAME}"

FallbackValue = Union[Any, Callable[[], Any]]

CAMEL_TO_SNAKE_PATTERN = re.compile(r"(?<=[a-z0-9])(?=[A-Z])")
NON_WORD_PATTERN = re.compile(r"[^0-9a-zA-Z_]+")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Fetch recent entries from the Boom Guru test table and issue POST requests "
            "to the configured inference endpoint."
        )
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=10,
        help="Number of rows to replay. Default: 10.",
    )
    parser.add_argument(
        "--order-column",
        dest="order_column",
        type=str,
        default=None,
        help=(
            "Column to use when determining recency. If omitted, the script attempts "
            "to pick a sensible default."
        ),
    )
    parser.add_argument(
        "--endpoint",
        type=str,
        default=DEFAULT_ENDPOINT,
        help=f"HTTP endpoint that will receive the replayed requests. Default: {DEFAULT_ENDPOINT}.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=(
            "Path of the JSON file that will contain the responses. Default: "
            f"{DEFAULT_OUTPUT_PATH}."
        ),
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=30.0,
        help="HTTP request timeout in seconds. Default: 30s.",
    )
    parser.add_argument(
        "--default-language",
        type=str,
        default="tr",
        help="Fallback language value when the database row does not provide one.",
    )
    parser.add_argument(
        "--default-webhook-url",
        type=str,
        default="http://localhost:8092/webhook-receiver",
        help="Fallback webhook URL when the database row does not provide one.",
    )
    parser.add_argument(
        "--default-form-id",
        type=str,
        default=None,
        help="Fallback form_id when the database row does not provide one.",
    )
    parser.add_argument(
        "--default-question-id",
        type=str,
        default=None,
        help="Fallback question_id when the database row does not provide one.",
    )
    parser.add_argument(
        "--default-image-url",
        type=str,
        default=None,
        help="Fallback image_url when the database row does not provide one.",
    )
    parser.add_argument(
        "--default-image-id",
        type=str,
        default=None,
        help="Fallback image_id when the database row does not provide one.",
    )
    parser.add_argument(
        "--default-serial-number",
        type=str,
        default=None,
        help="Fallback serial_number when the database row does not provide one.",
    )
    manual_source = parser.add_mutually_exclusive_group()
    manual_source.add_argument(
        "--image-url",
        type=str,
        default=None,
        help=(
            "If provided, skip the database lookup and send a single request using this image "
            "URL."
        ),
    )
    manual_source.add_argument(
        "--xlsx-path",
        type=Path,
        default=None,
        help=(
            "If provided, skip the database lookup and load rows from the specified Excel file "
            "instead (for example, a file under tests/files)."
        ),
    )
    parser.add_argument(
        "--xlsx-sheet",
        type=str,
        default=None,
        help=(
            "When using --xlsx-path, select the worksheet to read. Default: the active sheet."
        ),
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Fetch rows and show the payloads without issuing HTTP requests.",
    )
    args = parser.parse_args()
    if args.xlsx_sheet and not args.xlsx_path:
        parser.error("--xlsx-sheet can only be used together with --xlsx-path")
    return args

def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    args = parse_args()
    if args.xlsx_path:
        responses = run_from_excel(args)
    elif args.image_url:
        LOGGER.info("Using provided image URL; skipping database lookup.")
        responses = [
            launch(
                args.image_url,
                endpoint=args.endpoint,
                language=args.default_language,
                webhook_url=args.default_webhook_url,
                form_id=args.default_form_id,
                question_id=args.default_question_id,
                image_id=args.default_image_id,
                serial_number=args.default_serial_number,
                dry_run=args.dry_run,
                timeout=args.timeout,
            )
        ]
    else:
        with get_db_connection() as connection:
            cursor = connection.cursor()
            available_columns = set(fetch_column_names(cursor))
            order_column = choose_order_column(args.order_column, available_columns)
            rows = fetch_recent_rows(cursor, limit=args.limit, order_column=order_column)

        LOGGER.info("Fetched %d row(s) from %s", len(rows), FULL_TABLE_NAME)
        fallback_values = create_fallback_values(
            language=args.default_language,
            webhook_url=args.default_webhook_url,
            form_id=args.default_form_id,
            question_id=args.default_question_id,
            image_url=args.default_image_url,
            image_id=args.default_image_id,
            serial_number=args.default_serial_number,
        )
        responses = replay_requests(
            rows,
            fallback_values=fallback_values,
            endpoint=args.endpoint,
            timeout=args.timeout,
            dry_run=args.dry_run,
        )

    write_responses(args.output, responses)
    LOGGER.info("Stored %d response(s) in %s", len(responses), args.output)


def run_from_excel(args: argparse.Namespace) -> list[dict[str, Any]]:
    if args.xlsx_path is None:
        raise ValueError("--xlsx-path must be provided when running from Excel")

    try:
        rows = load_rows_from_xlsx(args.xlsx_path, sheet_name=args.xlsx_sheet)
    except (FileNotFoundError, ValueError, RuntimeError) as exc:
        LOGGER.error("Unable to load Excel data: %s", exc)
        raise SystemExit(1) from exc
    if not rows:
        LOGGER.warning("No usable rows found in Excel file %s", args.xlsx_path)
    else:
        LOGGER.info(
            "Loaded %d row(s) from Excel file %s%s",
            len(rows),
            args.xlsx_path,
            f" (sheet: {args.xlsx_sheet})" if args.xlsx_sheet else "",
        )

    fallback_values = create_fallback_values(
        language=args.default_language,
        webhook_url=args.default_webhook_url,
        form_id=args.default_form_id,
        question_id=args.default_question_id,
        image_url=args.default_image_url,
        image_id=args.default_image_id,
        serial_number=args.default_serial_number,
    )

    return replay_requests(
        rows,
        fallback_values=fallback_values,
        endpoint=args.endpoint,
        timeout=args.timeout,
        dry_run=args.dry_run,
    )


def fetch_column_names(cursor: Any) -> list[str]:
    """Return the column names for the configured test table."""
    query = (
        "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
        "WHERE TABLE_CATALOG = ? AND TABLE_SCHEMA = ? AND TABLE_NAME = ?"
    )
    cursor.execute(query, (MSSQL_DATABASE, TABLE_SCHEMA, TABLE_NAME))
    return [row[0] for row in cursor.fetchall()]


def choose_order_column(override: str | None, available_columns: Iterable[str]) -> str | None:
    normalized = {column.lower(): column for column in available_columns}
    if override:
        cleaned = override.strip()
        validate_column_name(cleaned)
        try:
            return normalized[cleaned.lower()]
        except KeyError as exc:
            raise ValueError(
                f"Column '{cleaned}' does not exist on {FULL_TABLE_NAME}."
            ) from exc

    for candidate in DEFAULT_ORDER_CANDIDATES:
        if candidate.lower() in normalized:
            return normalized[candidate.lower()]

    LOGGER.warning(
        "Could not determine an order column automatically; the query will not include ORDER BY."
    )
    return None


def validate_column_name(column: str) -> None:
    if not re.fullmatch(r"[A-Za-z0-9_]+", column):
        raise ValueError(f"Column name '{column}' contains invalid characters.")


def fetch_recent_rows(cursor: Any, *, limit: int, order_column: str | None) -> list[dict[str, Any]]:
    if limit <= 0:
        raise ValueError("Limit must be a positive integer.")

    order_clause = f" ORDER BY {order_column} DESC" if order_column else ""
    query = (
        f"SELECT TOP ({limit}) * FROM {FULL_TABLE_NAME} "
        "WHERE serial_number NOT LIKE ('SRC%')"
        f"{order_clause}"
    )
    cursor.execute(query)
    columns = [description[0] for description in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def create_fallback_values(
    *,
    language: FallbackValue | None,
    webhook_url: FallbackValue | None,
    form_id: FallbackValue | None,
    question_id: FallbackValue | None,
    image_url: FallbackValue | None,
    image_id: FallbackValue | None,
    serial_number: FallbackValue | None,
) -> dict[str, FallbackValue]:
    normalized_language = _sanitize_fallback(language)
    normalized_webhook = _sanitize_fallback(webhook_url)
    normalized_form_id = _sanitize_fallback(form_id)
    normalized_question_id = _sanitize_fallback(question_id)
    normalized_image_url = _sanitize_fallback(image_url)
    normalized_image_id = _sanitize_fallback(image_id)
    normalized_serial = _sanitize_fallback(serial_number)

    return {
        "language": normalized_language,
        "webhook_url": (normalized_webhook or DEFAULT_WEBHOOK_URL),
        "form_id": normalized_form_id,
        "question_id": normalized_question_id,
        "image_url": normalized_image_url,
        "image_id": normalized_image_id or generate_default_image_id(),
        "serial_number": normalized_serial or generate_default_serial_number(),
    }


def _sanitize_fallback(value: FallbackValue | None) -> FallbackValue | None:
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def resolve_fallback(value: FallbackValue | None) -> Any:
    if value is None:
        return None
    return value() if callable(value) else value


def load_rows_from_xlsx(path: Path, *, sheet_name: str | None = None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - optional dependency guard
        raise RuntimeError(
            "openpyxl is required to load Excel files. Install it with 'pip install openpyxl'."
        ) from exc

    if not path.exists():
        raise FileNotFoundError(f"Excel file '{path}' does not exist.")

    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise ValueError(
                f"Sheet '{sheet_name}' was not found in {path}. Available sheets: {available}"
            )
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active

    try:
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            return []

        normalized_headers: list[str | None] = []
        seen_headers: set[str] = set()
        for header in headers:
            if header is None:
                normalized_headers.append(None)
                continue
            normalized = normalize_excel_header(str(header))
            if not normalized:
                normalized_headers.append(None)
                continue
            if normalized in seen_headers:
                raise ValueError(
                    f"Duplicate column '{normalized}' detected in Excel file {path}."
                )
            seen_headers.add(normalized)
            normalized_headers.append(normalized)

        rows: list[dict[str, Any]] = []
        for raw_row in rows_iter:
            row_data: dict[str, Any] = {}
            contains_value = False
            for header, cell_value in zip(normalized_headers, raw_row):
                if header is None:
                    continue
                if cell_value is None:
                    continue
                if isinstance(cell_value, str):
                    stripped = cell_value.strip()
                    if stripped == "":
                        continue
                    row_data[header] = stripped
                else:
                    row_data[header] = cell_value
                if row_data[header] not in (None, ""):
                    contains_value = True
            if contains_value:
                rows.append(row_data)

        return rows
    finally:
        workbook.close()


def normalize_excel_header(raw_header: str) -> str:
    header = CAMEL_TO_SNAKE_PATTERN.sub("_", raw_header.strip())
    header = re.sub(r"[\s\-]+", "_", header)
    header = NON_WORD_PATTERN.sub("", header)
    header = re.sub(r"__+", "_", header)
    return header.strip("_").lower()


def build_payload(row: dict[str, Any], fallback_values: dict[str, FallbackValue]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    missing_fields: list[str] = []
    for field in REQUIRED_FIELDS:
        value = row.get(field)
        if value in (None, ""):
            value = fallback_values.get(field)
        if value in (None, ""):
            value = resolve_fallback(fallback_values.get(field))
        else:
            payload[field] = value

    for field in OPTIONAL_FIELDS:
        value = row.get(field)
        if value in (None, ""):
            value = resolve_fallback(fallback_values.get(field))
        if value is not None:
            payload[field] = value

    if missing_fields:
        raise ValueError(
            f"Row for serial_number={row.get('serial_number')} is missing required fields: {', '.join(missing_fields)}"
        )

    if not "image_url" in payload or not payload["image_url"]:
        payload["image_url"] = row.get("image_url") 

    return payload


def replay_requests(
    rows: Sequence[dict[str, Any]],
    *,
    fallback_values: dict[str, FallbackValue],
    endpoint: str,
    timeout: float,
    dry_run: bool,
) -> list[dict[str, Any]]:
    responses: list[dict[str, Any]] = []
    total = len(rows)
    for index, row in enumerate(rows, start=1):
        payload = build_payload(row, fallback_values)
        LOGGER.info(
            "[%d/%d] Sending request for serial_number=%s", index, total, payload.get("serial_number")
        )
        if dry_run:
            responses.append(
                {
                    "serial_number": payload.get("serial_number"),
                    "request_payload": payload,
                    "response": None,
                    "status_code": None,
                }
            )
            continue

        try:
            webhook_fallback = resolve_fallback(fallback_values.get("webhook_url"))
            if webhook_fallback is not None:
                payload.setdefault("webhook_url", webhook_fallback)
            response = requests.post(
                endpoint,
                json=payload,
                timeout=timeout,
            )
            response.raise_for_status()
            parsed_body = try_parse_json(response)
            status = response.status_code
        except requests.HTTPError as http_error:
            parsed_body = safe_error_payload(http_error.response)
            status = http_error.response.status_code if http_error.response else None
            LOGGER.error(
                "Request for serial_number=%s failed with status %s",
                payload.get("serial_number"),
                status,
            )
        except requests.RequestException as request_error:
            parsed_body = {"error": str(request_error)}
            status = None
            LOGGER.error(
                "Request for serial_number=%s failed: %s",
                payload.get("serial_number"),
                request_error,
            )

        responses.append(
            {
                "serial_number": payload.get("serial_number"),
                "request_payload": payload,
                "response": parsed_body,
                "status_code": status,
            }
        )

    return responses


def launch(
    image_url: str,
    *,
    endpoint: str = DEFAULT_ENDPOINT,
    timeout: float = 30.0,
    language: str = "tr",
    webhook_url: str | None = DEFAULT_WEBHOOK_URL,
    form_id: str | None = None,
    question_id: str | None = None,
    image_id: str | None = None,
    serial_number: str | None = None,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Send a single manual request using the provided image URL."""

    row: dict[str, Any] = {"image_url": image_url}
    if webhook_url is not None:
        row["webhook_url"] = webhook_url
    if language is not None:
        row["language"] = language
    if form_id is not None:
        row["form_id"] = form_id
    if question_id is not None:
        row["question_id"] = question_id
    if image_id is not None:
        row["image_id"] = image_id
    if serial_number is not None:
        row["serial_number"] = serial_number

    fallback_values = create_fallback_values(
        language=language,
        webhook_url=webhook_url,
        form_id=form_id,
        question_id=question_id,
        image_url=image_url,
        image_id=image_id,
        serial_number=serial_number,
    )

    responses = replay_requests(
        [row],
        fallback_values=fallback_values,
        endpoint=endpoint,
        timeout=timeout,
        dry_run=dry_run,
    )
    return responses[0]


def try_parse_json(response: requests.Response) -> Any:
    try:
        return response.json()
    except ValueError:
        return {"raw": response.text}


def safe_error_payload(response: requests.Response | None) -> dict[str, Any]:
    if response is None:
        return {"error": "No HTTP response returned."}
    try:
        return response.json()
    except ValueError:
        return {"status_code": response.status_code, "raw": response.text}


def write_responses(path: Path, responses: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    serializable = json.loads(json.dumps(responses, default=_json_default))
    path.write_text(json.dumps(serializable, ensure_ascii=False, indent=2))


def _json_default(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return str(value)

def generate_default_image_id() -> str:
    return f"manual-image-{uuid4().hex}"


def generate_default_serial_number() -> str:
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    suffix = uuid4().hex[:6]
    return f"manual-{timestamp}-{suffix}"



if __name__ == "__main__":
    main()